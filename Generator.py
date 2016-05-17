from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import get_column_letter

from os import listdir

inputDataFlag = 0
stringTableFlag = 0


for nameFile in listdir("."):
    if nameFile == 'InputData.xlsx':
        inputDataFlag = 1
    if nameFile == 'StringTable.xlsx':
        stringTableFlag = 1

if inputDataFlag == 0:
    print('Отстутсвует файл InputData.xlsx')


if inputDataFlag == 1:
    inputFile = 'InputData.xlsx'
    wb = load_workbook(filename = inputFile)
    sheet = wb['Vars']

    # Заполнение списка структур
    col = 'H'
    structList = []
    row = 2
    cell = col + str(row)
    while str(sheet[cell].value) != 'None' and row < 1000:
    #    print(sheet[cell].value)
        structList.append(str(sheet[cell].value))
        row = row + 1
        cell = col + str(row)

    print (structList)

    # Заполнение списка названий экранов
    col = 'I'
    screenList = []
    row = 2
    cell = col + str(row)
    while str(sheet[cell].value) != 'None' and row < 1000:
    #    print(sheet[cell].value)
        screenList.append(str(sheet[cell].value))
        row = row + 1
        cell = col + str(row)

    print (screenList)

    startLW = sheet['F1'].value
    startLB = sheet['F2'].value
    screenID = sheet['F3'].value
    varPrefix = sheet['F4'].value
    screenNameVar = sheet['F5'].value
    PLCName = sheet['F6'].value
    StringTable_ID = sheet['F7'].value

    # Заполняем матрицу переменных
    varRow = []
    varMatrix = []
    row = 2
    col = 1
    cell = get_column_letter(col) + str(row)
    while str(sheet[cell].value) != 'None' and row < 1000:
        for col in range(1, 4):
            cell = get_column_letter(col) + str(row)
            varRow.append(sheet[cell].value)
        row += 1
        cell = get_column_letter(col) + str(row)
        varMatrix.append(varRow)
    #    print(varRow)
        varRow = []

    print(varMatrix)

    index = 0
    for typeVar in varMatrix:
        varMatrix[index][1] = typeVar[1].upper()
        index += 1
    print (varMatrix)


    # Определяем типы переменных
    index = 0
    for variable in varMatrix:
        if (variable[1] == 'LB') or (variable[1] == 'BOOL'):
            varMatrix[index][1] = 1  #1 - тип Бит
        if (variable[1] == 'WORD') or (variable[1] == 'INT') or (variable[1] == 'LW') or (variable[1] == 'INTEGER'):
            varMatrix[index][1] = 2  #2 - тип слово
        if (variable[1] == 'REAL') or (variable[1] == 'DINT') or (variable[1] == 'DWORD'):
            varMatrix[index][1] = 3  #3 - тип двойное слово
        index += 1
    print(varMatrix)


    # Создаем книгу
    wb = Workbook()
    dest_filename = 'OutputVar.xlsx'
    ws = wb.active
    ws.title = 'HMI_Vars'

    row = 1
    col = 1
    lwIndex = startLW
    lbIndex = startLB

    for nameVar in varMatrix:
        cell = get_column_letter(col) + str(row)
        ws[cell] = varPrefix + 'Old_' + nameVar[0]
        cell = get_column_letter(col + 1) + str(row)
        ws[cell] = 'Local HMI'

        cell = get_column_letter(col + 2) + str(row)
        if nameVar[1] == 1:
            ws[cell] = 'LB'
            cell = get_column_letter(col + 3) + str(row)
            ws[cell] = str(lbIndex)
            lbIndex += 1
        if nameVar[1] == 2:
            ws[cell] = 'LW'
            cell = get_column_letter(col + 3) + str(row)
            ws[cell] = str(lwIndex)
            lwIndex += 1
        if nameVar[1] == 3:
            ws[cell] = 'LW'
            cell = get_column_letter(col + 3) + str(row)
            ws[cell] = str(lwIndex)
            lwIndex += 2

        cell = get_column_letter(col + 4) + str(row)
        ws[cell] = ''
        cell = get_column_letter(col + 5) + str(row)
        ws[cell] = 'Неопределенный'
        row += 1

        wb.save(filename = dest_filename)



    startScript = """
    macro_command main()

    short LcParameterSetNumber, TempShort, TempOldShort, ConstZero
    unsigned int TempInt, TempOldInt, ConstIntZero
    bool ConstTrue, ConstFalse, TempBool, TempOldBool

    ConstTrue  = true
    ConstFalse = false
    ConstZero  = 0

    """
    str1 = 'GetData(LcParameterSetNumber, "Local HMI", "' + screenID + '", 1)'

    codeFileName = 'OutputCode.txt'
    outputFile = open(codeFileName,'w')

    outputFile.write(startScript+str1+'\n')

    currStructNumber = 0
    endStructs = len(structList)
    for j in structList:
        currNumVar = 0
        if currStructNumber == 0:
            str1 = 'if LcParameterSetNumber == ' + str(currStructNumber + 1) + ' then\n'
            print(str1)
            #outputFile.write(str1+'\n')
        else:
            str1 = 'else if LcParameterSetNumber == ' + str(currStructNumber + 1) + ' then\n'
            print(str1)

        #str2 = '    StringCopy("' + screenList[currStructNumber] + '", ScreenName[0])\n'
        #str3 = '    StringSet(ScreenName[0], "Local HMI", "' + screenNameVar + '", 20)\n'
        outputFile.write(str1 + '\n')
        for i in varMatrix:
            if i[2] == 1:
                if i[1] == 1:
                    str1 = '    GetData(TempOldBool, "Local HMI", "' + varPrefix + 'Old_' + i[0] + '", 1)\n'
                    str2 = '    GetData(TempBool, "Local HMI", "' + varPrefix + i[0] + '", 1)\n'
                    str3 = '    if 	(TempOldBool <> TempBool) then\n'
                    str4 = '        SetData(TempBool, "' + PLCName + '", "' + structList[currStructNumber] + '.' + i[0] + '", 1)\n'
                    str5 = '    else\n'
                    str6 = '        GetData(TempBool, "' + PLCName + '", "' + structList[currStructNumber] + '.' + i[0] + '", 1)\n'
                    str7 = '        SetData(TempBool, "Local HMI", "' + varPrefix + i[0] + '", 1)\n'
                    str8 = '    end if\n'
                    str9 = '    GetData(TempBool, "Local HMI", "' + varPrefix + i[0] + '", 1)\n'
                    str10 = '    SetData(TempBool, "Local HMI", "' + varPrefix + 'Old_' + i[0] + '", 1)\n'
                if i[1] == 2:
                    str1 = '    GetData(TempOldShort, "Local HMI", "' + varPrefix + 'Old_' + i[0] + '", 1)\n'
                    str2 = '    GetData(TempShort, "Local HMI", "' + varPrefix + i[0] + '", 1)\n'
                    str3 = '    if 	(TempOldShort <> TempShort) then\n'
                    str4 = '        SetData(TempShort, "' + PLCName + '", "' + structList[currStructNumber] + '.' + i[0] + '", 1)\n'
                    str5 = '    else\n'
                    str6 = '        GetData(TempShort, "' + PLCName + '", "' + structList[currStructNumber] + '.' + i[0] + '", 1)\n'
                    str7 = '        SetData(TempShort, "Local HMI", "' + varPrefix + i[0] + '", 1)\n'
                    str8 = '    end if\n'
                    str9 = '    GetData(TempShort, "Local HMI", "' + varPrefix + i[0] + '", 1)\n'
                    str10 = '    SetData(TempShort, "Local HMI", "' + varPrefix + 'Old_' + i[0] + '", 1)\n'
                if i[1] == 3:
                    str1 = '    GetData(TempOldInt, "Local HMI", "'+ varPrefix + 'Old_' + i[0] + '", 2)\n'
                    str2 = '    GetData(TempInt, "Local HMI", "'+ varPrefix + i[0] + '", 2)\n'
                    str3 = '    if 	(TempOldInt <> TempInt) then\n'
                    str4 = '        SetData(TempInt, "' + PLCName + '", "'+ structList[currStructNumber] + '.' + i[0] + '", 2)\n'
                    str5 = '    else\n'
                    str6 = '        GetData(TempInt, "' + PLCName + '", "'+ structList[currStructNumber] + '.' + i[0] + '", 2)\n'
                    str7 = '        SetData(TempInt, "Local HMI", "'+ varPrefix + i[0] + '", 2)\n'
                    str8 = '    end if\n'
                    str9 = '    GetData(TempInt, "Local HMI", "'+ varPrefix + i[0] + '", 2)\n'
                    str10 = '    SetData(TempInt, "Local HMI", "'+ varPrefix + 'Old_' + i[0] + '", 2)\n'
                print(str1+str2+str3+str4+str5+str6+str7+str8+str9+str10)
                outputFile.write(str1+str2+str3+str4+str5+str6+str7+str8+str9+str10+'\n')
            else:
                if i[1] == 1:
                    str1 = '    GetData(TempBool, "' + PLCName + '", "' + structList[currStructNumber] + '.' + i[0] + '", 1)\n'
                    str2 = '    SetData(TempBool, "Local HMI", "' + varPrefix + i[0] + '", 1)\n'
                if i[1] == 2:
                    str1 = '    GetData(TempShort, "' + PLCName + '", "' + structList[currStructNumber] + '.' + i[0] + '", 1)\n'
                    str2 = '    SetData(TempShort, "Local HMI", "' + varPrefix + i[0] + '", 1)\n'
                if i[1] == 3:
                    str1 = '    GetData(TempInt, "' + PLCName + '", "' + structList[currStructNumber] + '.' + i[0] + '", 2)\n'
                    str2 = '    SetData(TempInt, "Local HMI", "' + varPrefix + i[0] + '", 2)\n'
                print(str1+str2)
                outputFile.write(str1+str2+'\n')
            print(i)
            currNumVar += 1
        currStructNumber += 1
        if currStructNumber == endStructs:
            str1 = 'end if\n'
            str2 = '\n'
            str3 = 'end macro_command\n'
            print(str1+str2+str3)
            outputFile.write(str1+str2+str3+'\n')

    outputFile.close()

    inputFile = 'StringTable.xlsx'
    try:
        wb = load_workbook(filename = inputFile)
        ws = wb.active
    except:
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'ID раздела'
        ws['B1'] = 'Описание'
        ws['C1'] = 'Номер строки'
        ws['D1'] = 'язык 1'
        ws['E1'] = 'язык 2'
        ws['F1'] = 'язык 3'
        ws['G1'] = 'язык 4'
        ws['H1'] = 'язык 5'
        ws['I1'] = 'язык 6'
        ws['J1'] = 'язык 7'
        ws['K1'] = 'язык 8'

    row = 2
    col = 1
    cell = get_column_letter(col) + str(row)
    while str(ws[cell].value) != 'None' and row < 1000:
    #    print(sheet[cell].value)
        row = row + 1
        cell = get_column_letter(col) + str(row)


    col = 1
    index = 0
    for screenName in screenList:
        cell = get_column_letter(col) + str(row)
        ws[cell] = str(StringTable_ID)
        cell = get_column_letter(col + 1) + str(row)
        ws[cell] = varPrefix.replace('.','')
        cell = get_column_letter(col + 2) + str(row)
        ws[cell] = str(index)
        index += 1
        cell = get_column_letter(col + 3) + str(row)
        ws[cell] = str(screenName)
        row += 1

    wb.save(filename = inputFile)

    print('Создан файл импорта переменных в Weintek "' + dest_filename + '"')
    print('Создан файл макроса для Weintek "' + codeFileName + '"')
    print('Создан файл импорта таблиц строк в Weintek "' + inputFile + '"')
    print('\n')




input('Press any key...')
