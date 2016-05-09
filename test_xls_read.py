screenName = 'Sensor'
prefixOld = 'old'


from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import get_column_letter
wb = load_workbook(filename = 'Var.xlsx')
sheet = wb['Structs']

# Заполнение списка структур
col = 'A'
structList = []
row = 2
cell = col + str(row)
while str(sheet[cell].value) != 'None' and row < 1000:
#    print(sheet[cell].value)
    structList.append(str(sheet[cell].value))
    row = row + 1
    cell = col + str(row)

print (structList)

# Переходим на другой лист
sheet = wb['Vars']

# Заполняем матрицу переменных
varRow = []
varMatrix = []
row = 2
col = 1
cell = get_column_letter(col) + str(row)
while str(sheet[cell].value) != 'None' and row < 1000:
    for col in range(1, 4):
        cell = get_column_letter(col) + str(row)
        varRow.append(sheet[cell].value)
    row += 1
    cell = get_column_letter(col) + str(row)
    varMatrix.append(varRow)
#    print(varRow)
    varRow = []

print(varMatrix)


# Открываем файл экспорта из Weintek
wb = load_workbook(filename = 'test.xlsx')
ws = wb.active

# Находим свободную строку
col = 'A'
row = 1
cell = col + str(row)
while str(ws[cell].value) != 'None' and row < 10000:
#    print(ws[cell].value)
    row += 1
    cell = col + str(row)

emptyRow = row

row = emptyRow
col = 1
memoryIndex = 4000
for nameVar in varMatrix:
    cell = get_column_letter(col) + str(row)
    ws[cell] = nameVar[0]
    cell = get_column_letter(col + 1) + str(row)
    ws[cell] = 'Local HMI'
    cell = get_column_letter(col + 2) + str(row)
    ws[cell] = 'LW'
    cell = get_column_letter(col + 3) + str(row)
    ws[cell] = str(memoryIndex - emptyRow + row)
    cell = get_column_letter(col + 4) + str(row)
    ws[cell] = ''
    cell = get_column_letter(col + 5) + str(row)
    ws[cell] = 'Неопределенный'
    row += 1
#cell = get_column_letter(col) + str(row)
#print(cell)
#ws[cell] = 'Мама мыла раму'

#wb = Workbook()
dest_filename = 'outputVars.xlsx'

#ws = wb.active
ws.title = 'HMI_Vars'



wb.save(filename = dest_filename)

#input('press enter any key...')